#!/usr/bin/env python3
"""
Conferência de plantões médicos — HIAE Goiânia
Uso: python3 conferir.py <arquivo_escala.xls[x]> [saida.xlsx]

Lê a escala exportada e gera um Excel com:
  - Aba "Resumo": contagem por médico e categoria (com fórmulas)
  - Aba "Detalhado": lista completa de plantões
  - Aba "Noturnos": lista de noturnos para preenchimento manual de "Acionado? (S)"
    - Deixar em branco = não acionado. Marcar "S" = acionado.
"""

import sys
import os
import re
import shutil
import tempfile
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Utilidades ────────────────────────────────────────────────
LOWER_WORDS = {'De', 'Da', 'Do', 'Das', 'Dos', 'E'}

def normalize_name(name):
    if not name: return None
    name = ' '.join(str(name).split())
    words = name.title().split()
    out = []
    for i, w in enumerate(words):
        if i > 0 and w in LOWER_WORDS:
            out.append(w.lower())
        else:
            out.append(w)
    return ' '.join(out)

def parse_date(s):
    if s is None: return None
    if isinstance(s, datetime): return s
    m = re.match(r'(\d{2})/(\d{2})/(\d{4})', str(s))
    if not m: return None
    d, mo, y = m.groups()
    return datetime(int(y), int(mo), int(d))

def detect_shift(cell):
    if not cell: return None
    s = str(cell).lower()
    if 'manhã' in s or 'manha' in s: return 'Manhã'
    if 'tarde' in s: return 'Tarde'
    if 'noite' in s or 'noturno' in s: return 'Noite'
    return None

DIAS_PT = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']

CAT_MEIO  = 'Meio de semana'
CAT_FDS   = 'Fim de semana'
CAT_NOITE = 'Noturno'

def categorize(p):
    """Categoria do plantão. Noturno é tratado separado (aba Noturnos)."""
    if p['turno'] == 'Noite':
        return CAT_NOITE
    dia = p['data'].weekday()  # 0=seg ... 6=dom
    return CAT_MEIO if dia < 5 else CAT_FDS

# ── Parser ────────────────────────────────────────────────────
def parse_escala(xlsx_path):
    """Retorna lista de plantões: [{data, dia_semana, fim_semana, turno, medico}, ...]"""
    # Alguns arquivos vêm com extensão .xls mas são .xlsx — copia como .xlsx
    tmp = None
    if not xlsx_path.lower().endswith('.xlsx'):
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        shutil.copy(xlsx_path, tmp.name)
        xlsx_path = tmp.name
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    plantoes = []
    i = 0
    while i < len(rows):
        # Procura cabeçalho de semana (linha que começa com None e tem "Segunda" em col B)
        row = rows[i]
        if row and len(row) >= 8 and row[1] == 'Segunda':
            # Próxima linha deve ter as datas
            if i + 1 < len(rows):
                date_row = rows[i + 1]
                datas = [parse_date(date_row[c]) for c in range(1, 8)]
                # A partir de i+2, lê linhas até próximo cabeçalho ou fim
                j = i + 2
                while j < len(rows):
                    r = rows[j]
                    if not r or all(v is None for v in r): break
                    turno = detect_shift(r[0])
                    if turno is None: break
                    for c in range(7):
                        medico = normalize_name(r[c + 1])
                        data = datas[c]
                        if medico and data:
                            dia_idx = data.weekday()  # 0=segunda ... 6=domingo
                            plantoes.append({
                                'data': data,
                                'dia_semana': DIAS_PT[dia_idx],
                                'fim_semana': dia_idx >= 5,  # sábado ou domingo
                                'turno': turno,
                                'medico': medico,
                            })
                    j += 1
                i = j
                continue
        i += 1
    if tmp:
        os.unlink(tmp.name)
    return plantoes

# ── Geração do Excel de saída ─────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
ALT_FILL = PatternFill('solid', fgColor='F2F2F2')
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
BORDER = Border(*(Side(border_style='thin', color='CCCCCC'),) * 4)

def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

def gerar_excel(plantoes, out_path, periodo_label):
    wb = Workbook()

    # ── Aba Detalhado ───────────────────────────────────
    ws_det = wb.active
    ws_det.title = 'Detalhado'
    headers_det = ['Data', 'Dia da semana', 'Turno', 'Categoria', 'Médico']
    for c, h in enumerate(headers_det, 1):
        cell = ws_det.cell(1, c, h)
        style_header(cell)
    for r, p in enumerate(sorted(plantoes, key=lambda x: (x['data'], x['turno'], x['medico'])), 2):
        ws_det.cell(r, 1, p['data'].strftime('%d/%m/%Y')).alignment = CENTER
        ws_det.cell(r, 2, p['dia_semana']).alignment = CENTER
        ws_det.cell(r, 3, p['turno']).alignment = CENTER
        ws_det.cell(r, 4, categorize(p)).alignment = CENTER
        ws_det.cell(r, 5, p['medico']).alignment = LEFT
        if r % 2 == 0:
            for c in range(1, 6):
                ws_det.cell(r, c).fill = ALT_FILL
    for c, w in enumerate([12, 14, 10, 22, 42], 1):
        ws_det.column_dimensions[get_column_letter(c)].width = w
    ws_det.freeze_panes = 'A2'

    # ── Aba Noturnos ────────────────────────────────────
    ws_not = wb.create_sheet('Noturnos')
    headers_not = ['Data', 'Dia da semana', 'Médico', 'Acionado? (S)']
    for c, h in enumerate(headers_not, 1):
        cell = ws_not.cell(1, c, h)
        style_header(cell)
    noturnos = [p for p in plantoes if p['turno'] == 'Noite']
    for r, p in enumerate(sorted(noturnos, key=lambda x: (x['data'], x['medico'])), 2):
        ws_not.cell(r, 1, p['data'].strftime('%d/%m/%Y')).alignment = CENTER
        ws_not.cell(r, 2, p['dia_semana']).alignment = CENTER
        ws_not.cell(r, 3, p['medico']).alignment = LEFT
        ws_not.cell(r, 4, '').alignment = CENTER
        if r % 2 == 0:
            for c in range(1, 5):
                ws_not.cell(r, c).fill = ALT_FILL
    for c, w in enumerate([12, 14, 42, 18], 1):
        ws_not.column_dimensions[get_column_letter(c)].width = w
    ws_not.freeze_panes = 'A2'

    # ── Aba Resumo ──────────────────────────────────────
    ws_res = wb.create_sheet('Resumo', 0)  # primeira aba
    ws_res.cell(1, 1, f'Conferência de plantões — {periodo_label}').font = Font(bold=True, size=14)
    ws_res.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws_res.cell(2, 1, 'Preencha "S" na aba Noturnos para plantões que foram acionados. O restante conta como não acionado.').font = Font(italic=True, color='666666', size=10)
    ws_res.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    header_row = 4
    headers_res = [
        'Médico',
        'Meio de semana', 'Fim de semana', 'Noturno não acionado', 'Noturno acionado', 'Total',
        'R$/h MS', 'R$ Meio de semana', 'R$ Fim de semana', 'R$ Not. não acionado', 'R$ Not. acionado', 'TOTAL R$',
    ]
    for c, h in enumerate(headers_res, 1):
        cell = ws_res.cell(header_row, c, h)
        style_header(cell)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    money_fmt = 'R$ #,##0.00'
    medicos = sorted(set(p['medico'] for p in plantoes))
    # Colunas: A médico | B MS | C FDS | D NnAc | E NAc | F Total | G R$/h | H R$MS | I R$FDS | J R$NnAc | K R$NAc | L Total R$
    for r_idx, medico in enumerate(medicos, header_row + 1):
        ws_res.cell(r_idx, 1, medico).alignment = LEFT
        # ── Contagens ────────────────────────────
        ws_res.cell(r_idx, 2, f'=COUNTIFS(Detalhado!E:E,A{r_idx},Detalhado!D:D,"{CAT_MEIO}")').alignment = CENTER
        ws_res.cell(r_idx, 3, f'=COUNTIFS(Detalhado!E:E,A{r_idx},Detalhado!D:D,"{CAT_FDS}")').alignment = CENTER
        ws_res.cell(r_idx, 5, f'=COUNTIFS(Noturnos!C:C,A{r_idx},Noturnos!D:D,"S")').alignment = CENTER
        ws_res.cell(r_idx, 4, f'=COUNTIF(Noturnos!C:C,A{r_idx})-E{r_idx}').alignment = CENTER
        ws_res.cell(r_idx, 6, f'=SUM(B{r_idx}:E{r_idx})').alignment = CENTER
        # ── Valores ──────────────────────────────
        # R$/h meio de semana (tabela progressiva pelo total de MS em B)
        ws_res.cell(r_idx, 7, f'=IF(B{r_idx}<=6,135,IF(B{r_idx}<=10,145,IF(B{r_idx}<=15,155,IF(B{r_idx}<=19,165,175))))').alignment = CENTER
        # R$ meio de semana = qtde * 6h * R$/h
        ws_res.cell(r_idx, 8, f'=B{r_idx}*6*G{r_idx}').number_format = money_fmt
        # R$ fim de semana = qtde * 6h * R$165
        ws_res.cell(r_idx, 9, f'=C{r_idx}*6*165').number_format = money_fmt
        # R$ noturno não acionado = qtde * R$522
        ws_res.cell(r_idx, 10, f'=D{r_idx}*522').number_format = money_fmt
        # R$ noturno acionado = qtde * 12h * R$145
        ws_res.cell(r_idx, 11, f'=E{r_idx}*12*145').number_format = money_fmt
        # TOTAL R$
        ws_res.cell(r_idx, 12, f'=SUM(H{r_idx}:K{r_idx})').number_format = money_fmt
        ws_res.cell(r_idx, 12).font = Font(bold=True)
        if (r_idx - header_row) % 2 == 0:
            for c in range(1, 13):
                ws_res.cell(r_idx, c).fill = ALT_FILL

    # Linha de totais
    tot_row = header_row + len(medicos) + 1
    ws_res.cell(tot_row, 1, 'TOTAL').font = Font(bold=True)
    ws_res.cell(tot_row, 1).alignment = LEFT
    for c in list(range(2, 7)) + list(range(8, 13)):
        col = get_column_letter(c)
        cell = ws_res.cell(tot_row, c, f'=SUM({col}{header_row+1}:{col}{tot_row-1})')
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        if c >= 8:
            cell.number_format = money_fmt
    for c in range(1, 13):
        ws_res.cell(tot_row, c).fill = PatternFill('solid', fgColor='DDEBF7')

    for c, w in enumerate([42, 15, 15, 20, 17, 8, 8, 18, 18, 20, 17, 15], 1):
        ws_res.column_dimensions[get_column_letter(c)].width = w
    ws_res.freeze_panes = 'B5'

    wb.save(out_path)

# ── Main ──────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    if not os.path.exists(src):
        print(f'Arquivo não encontrado: {src}')
        sys.exit(1)
    plantoes = parse_escala(src)
    if not plantoes:
        print('Nenhum plantão encontrado no arquivo. Verifique o formato.')
        sys.exit(1)
    # Detecta o mês de referência: tenta pelo nome do arquivo (ex: 01-07-2026),
    # senão usa o mês mais frequente entre os plantões.
    ref_year = ref_month = None
    m = re.search(r'(\d{2})-(\d{2})-(\d{4})', os.path.basename(src))
    if m:
        _, ref_month, ref_year = m.groups()
        ref_month, ref_year = int(ref_month), int(ref_year)
    else:
        from collections import Counter
        cnt = Counter((p['data'].year, p['data'].month) for p in plantoes)
        ref_year, ref_month = cnt.most_common(1)[0][0]
    # Filtra somente plantões do mês de referência
    plantoes = [p for p in plantoes if p['data'].year == ref_year and p['data'].month == ref_month]
    if not plantoes:
        print('Nenhum plantão encontrado no mês de referência.')
        sys.exit(1)
    meses_pt = ['','janeiro','fevereiro','março','abril','maio','junho','julho','agosto','setembro','outubro','novembro','dezembro']
    periodo = f"{meses_pt[ref_month].capitalize()}/{ref_year}"
    # Nome de saída: se não passado, deriva do input
    if len(sys.argv) >= 3:
        out = sys.argv[2]
    else:
        base = os.path.splitext(os.path.basename(src))[0]
        out = os.path.join(os.path.dirname(os.path.abspath(src)), f'{base}_conferencia.xlsx')
    gerar_excel(plantoes, out, periodo)
    print(f'✓ {len(plantoes)} plantões processados ({periodo})')
    print(f'✓ Excel gerado: {out}')

if __name__ == '__main__':
    main()
